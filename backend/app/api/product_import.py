"""
Product Import API — InvenIQ
Loads product files (CSV / XLSX) from the product_imp_files/ directory.
Provides search, upload, and AI-assisted matching endpoints used by the
Quotation Builder to lookup products by item code, name, or category.
"""
import csv
import io
import json
import logging
import os
import re
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile

logger = logging.getLogger(__name__)
router = APIRouter(tags=["Product Import"])

# ── Directory where product files live ───────────────────────────────────────
_FILES_DIR = Path(__file__).parent.parent.parent.parent / "product_imp_files"
_FILES_DIR.mkdir(exist_ok=True)

# ── In-memory store: {filename: [{"item_code","item_name","category",...}]} ──
_FILE_CACHE: dict[str, list[dict]] = {}
_LOADED = False

# ── Column aliases used across different supplier file formats ─────────────────
_COL_ALIASES = {
    "item_code":   ["item code", "item_code", "code", "sku", "part no", "part number", "product code", "item no", "article no", "article code", "model no", "model number"],
    "item_name":   ["item name", "item_name", "name", "description", "product name", "product", "article name", "model name", "title", "item description"],
    "category":    ["category", "cat", "product category", "group", "product group", "type", "product type"],
    "size":        ["size", "dimension", "dimensions", "specification", "specifications"],
    "finish":      ["finish", "color", "colour", "variant"],
    "unit":        ["unit", "uom", "pcs / set", "pcs/set", "pack", "packing"],
    "mrp":         ["mrp", "mrp (?)", "mrp (₹)", "price", "unit price", "rate", "sell price", "list price", "spu"],
    "brand":       ["brand", "make", "manufacturer", "company"],
    "hsn_code":    ["hsn", "hsn code", "hsn_code"],
}


def _normalise_key(col: str) -> str:
    c = col.strip().lower().replace("(₹)", "").replace("(?)", "").strip()
    for key, aliases in _COL_ALIASES.items():
        if c in aliases:
            return key
    return c.replace(" ", "_")


def _parse_csv_bytes(raw: bytes, filename: str) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    col_map = {f: _normalise_key(f) for f in reader.fieldnames}
    products = []
    for i, row in enumerate(reader):
        mapped = {col_map[k]: (v.strip() if v else "") for k, v in row.items()}
        # Ensure required fields
        if not mapped.get("item_code") and not mapped.get("item_name"):
            continue
        mapped.setdefault("item_code",  f"{filename[:4].upper()}-{i+1:04d}")
        mapped.setdefault("item_name",  "")
        mapped.setdefault("category",   "General")
        mapped.setdefault("brand",      "")
        mapped.setdefault("unit",       "Nos")
        mapped.setdefault("mrp",        "0")
        mapped.setdefault("size",       "")
        mapped.setdefault("finish",     "")
        mapped["_source_file"] = filename
        products.append(mapped)
    return products


def _parse_xlsx_bytes(raw: bytes, filename: str) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        products = []
        for sh_name in wb.sheetnames:
            ws = wb[sh_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # Find header row (first row with at least 2 non-None cells)
            header_row = None
            data_start = 0
            for ri, row in enumerate(rows):
                non_null = [c for c in row if c is not None]
                if len(non_null) >= 2:
                    header_row = [str(c).strip() if c is not None else "" for c in row]
                    data_start = ri + 1
                    break
            if not header_row:
                continue
            col_map = {i: _normalise_key(h) for i, h in enumerate(header_row)}
            for ri, row in enumerate(rows[data_start:]):
                mapped = {}
                for ci, val in enumerate(row):
                    key = col_map.get(ci, f"col_{ci}")
                    mapped[key] = str(val).strip() if val is not None else ""
                if not mapped.get("item_code") and not mapped.get("item_name"):
                    continue
                if all(not v for v in mapped.values()):
                    continue
                mapped.setdefault("item_code",  f"{sh_name[:4].upper()}-{ri+1:04d}")
                mapped.setdefault("item_name",  "")
                mapped.setdefault("category",   "General")
                mapped.setdefault("brand",      "")
                mapped.setdefault("unit",       "Nos")
                mapped.setdefault("mrp",        "0")
                mapped["_source_file"] = filename
                mapped["_sheet"]       = sh_name
                products.append(mapped)
        return products
    except Exception as exc:
        logger.warning("product_import: xlsx parse failed (%s) — %s", filename, exc)
        return []


def _load_file(path: Path) -> list[dict]:
    raw = path.read_bytes()
    fn  = path.name
    if fn.lower().endswith(".csv"):
        return _parse_csv_bytes(raw, fn)
    if fn.lower().endswith((".xlsx", ".xls")):
        return _parse_xlsx_bytes(raw, fn)
    return []


def _ensure_loaded():
    global _LOADED
    if _LOADED:
        return
    for p in sorted(_FILES_DIR.glob("*")):
        if p.name.startswith("~") or p.name.startswith("."):
            continue
        if p.suffix.lower() not in (".csv", ".xlsx", ".xls"):
            continue
        try:
            products = _load_file(p)
            if products:
                _FILE_CACHE[p.name] = products
                logger.info("product_import: loaded %d products from %s", len(products), p.name)
        except Exception as exc:
            logger.warning("product_import: failed to load %s — %s", p.name, exc)
    _LOADED = True


def _fuzzy_score(query: str, product: dict) -> int:
    """Return a relevance score 0-100. Higher = better match."""
    q = query.lower().strip()
    if not q:
        return 100
    score = 0
    # Exact item_code match
    code = (product.get("item_code") or "").lower()
    if q == code:
        return 100
    if q in code:
        score += 80
    # Item name match
    name = (product.get("item_name") or "").lower()
    if q in name:
        score += 60
    elif any(w in name for w in q.split()):
        score += 30
    # Category match
    cat = (product.get("category") or "").lower()
    if q in cat:
        score += 40
    # Size / finish match
    for fld in ("size", "finish", "brand"):
        if q in (product.get(fld) or "").lower():
            score += 10
    return min(score, 99)


def _search_in_list(products: list[dict], query: str, category: str, limit: int) -> list[dict]:
    results = []
    q_low = query.lower().strip()
    cat_low = category.lower().strip() if category else ""
    for p in products:
        if cat_low and cat_low not in (p.get("category") or "").lower():
            continue
        score = _fuzzy_score(q_low, p) if q_low else 50
        if score > 0:
            results.append({**p, "_score": score})
    results.sort(key=lambda x: x["_score"], reverse=True)
    return results[:limit]


# ── AI-assisted matching ───────────────────────────────────────────────────────

async def _ai_match(query: str, candidates: list[dict]) -> list[dict]:
    """Use GPT-4o to re-rank and enrich candidates with missing price/HSN data."""
    try:
        from app.core.config import get_settings
        cfg = get_settings()
        if not cfg.openai_api_key or not candidates:
            return candidates
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=cfg.openai_api_key, timeout=30.0)
        cand_text = json.dumps([
            {"item_code": c.get("item_code"), "item_name": c.get("item_name"),
             "category": c.get("category"), "size": c.get("size"), "mrp": c.get("mrp")}
            for c in candidates[:20]
        ], indent=None)
        prompt = f"""You are a hardware & building materials product expert for the Indian market.
Customer requirement: "{query}"
Candidate products from our catalog:
{cand_text}

Return JSON array of the top matching products (max 10), in relevance order.
For each, add fields: "hsn_code" (Indian HSN), "unit" (Nos/Set/Mtr/Sqft/Kg), "mrp_enriched" (number, use provided MRP or estimate).
Only return product item_codes from the provided list. No new products.
Return ONLY a JSON array, no text."""
        resp = await client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=600,
            response_format={"type": "json_object"},
        )
        raw = json.loads(resp.choices[0].message.content)
        ai_list = raw if isinstance(raw, list) else raw.get("products", raw.get("items", candidates))
        # Merge AI enrichments back into candidates
        cand_map = {c["item_code"]: c for c in candidates}
        enriched = []
        for ai_item in (ai_list or []):
            code = ai_item.get("item_code")
            if code in cand_map:
                merged = {**cand_map[code], **{k: v for k, v in ai_item.items() if v}}
                enriched.append(merged)
        return enriched or candidates
    except Exception as exc:
        logger.debug("product_import: AI match skipped — %s", exc)
        return candidates


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("/product-files")
async def list_product_files():
    """List all available product files with stats."""
    _ensure_loaded()
    files = []
    for fname, products in _FILE_CACHE.items():
        cats = list({p.get("category", "General") for p in products})
        files.append({
            "filename":   fname,
            "count":      len(products),
            "categories": sorted(cats)[:20],
            "columns":    sorted({k for p in products[:5] for k in p.keys() if not k.startswith("_")}),
        })
    # Also list files not yet loaded (empty or unparsed)
    for p in sorted(_FILES_DIR.glob("*")):
        if p.name.startswith("~") or p.name.startswith("."):
            continue
        if p.suffix.lower() not in (".csv", ".xlsx", ".xls"):
            continue
        if p.name not in _FILE_CACHE:
            files.append({"filename": p.name, "count": 0, "categories": [], "columns": []})
    return {"files": files, "total_products": sum(len(v) for v in _FILE_CACHE.values())}


@router.get("/product-files/search")
async def search_products(
    q:        str              = Query("",    description="Item code, name, or keyword"),
    category: str              = Query("",    description="Filter by category"),
    file:     Optional[str]    = Query(None,  description="Limit to specific filename"),
    limit:    int              = Query(20,    ge=1, le=200),
    ai:       bool             = Query(False, description="Use AI to re-rank and enrich results"),
):
    """Search across all loaded product files."""
    _ensure_loaded()
    results = []
    sources = {file: _FILE_CACHE[file]} if file and file in _FILE_CACHE else _FILE_CACHE
    for fname, products in sources.items():
        hits = _search_in_list(products, q, category, limit * 2)
        results.extend(hits)

    # Global sort + dedup by item_code
    results.sort(key=lambda x: x.get("_score", 0), reverse=True)
    seen = set()
    deduped = []
    for r in results:
        key = (r.get("item_code", ""), r.get("_source_file", ""))
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    deduped = deduped[:limit]

    if ai and q and deduped:
        deduped = await _ai_match(q, deduped)

    return {
        "results":   deduped,
        "count":     len(deduped),
        "query":     q,
        "category":  category,
        "ai_used":   ai and bool(q),
    }


@router.get("/product-files/categories")
async def list_categories(file: Optional[str] = Query(None)):
    """Return distinct categories across all files (or one file)."""
    _ensure_loaded()
    sources = {file: _FILE_CACHE[file]} if file and file in _FILE_CACHE else _FILE_CACHE
    cats: dict[str, int] = {}
    for products in sources.values():
        for p in products:
            c = p.get("category") or "General"
            cats[c] = cats.get(c, 0) + 1
    return {"categories": [{"name": k, "count": v} for k, v in sorted(cats.items(), key=lambda x: -x[1])]}


@router.get("/product-files/{filename}/products")
async def get_file_products(filename: str, limit: int = Query(100, le=1000), offset: int = 0):
    """Return all products from a specific file (paginated)."""
    _ensure_loaded()
    if filename not in _FILE_CACHE:
        raise HTTPException(404, f"File '{filename}' not found or not loaded")
    products = _FILE_CACHE[filename]
    return {
        "filename": filename,
        "products": products[offset: offset + limit],
        "total":    len(products),
        "offset":   offset,
        "limit":    limit,
    }


@router.post("/product-files/upload", status_code=201)
async def upload_product_file(file: UploadFile = File(...)):
    """Upload a new CSV or XLSX product file. Auto-loads into the search index."""
    global _LOADED
    if not file.filename:
        raise HTTPException(400, "filename is required")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".csv", ".xlsx", ".xls"):
        raise HTTPException(400, "Only CSV and XLSX files are supported")
    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(413, "File too large (max 20 MB)")
    dest = _FILES_DIR / file.filename
    dest.write_bytes(raw)
    # Parse and cache
    products = _load_file(dest)
    _FILE_CACHE[file.filename] = products
    logger.info("product_import: uploaded %s — %d products", file.filename, len(products))
    return {
        "filename": file.filename,
        "count":    len(products),
        "message":  f"Loaded {len(products)} products from {file.filename}",
    }


@router.post("/product-files/smart-match")
async def smart_match_products(payload: dict):
    """
    Match products using ANY combination of known fields.
    Accepts: item_code, category, size, finish, pcs_per_set, spu, mrp, mrp_min, mrp_max, query.
    Returns exact matches, strong matches, and near matches separately.
    Customers can describe products by ANY attribute — AI fills the rest.
    """
    _ensure_loaded()

    item_code   = (payload.get("item_code")   or "").strip().lower()
    category    = (payload.get("category")    or "").strip().lower()
    size        = (payload.get("size")        or "").strip().lower()
    finish      = (payload.get("finish")      or "").strip().lower()
    pcs_set     = (payload.get("pcs_per_set") or payload.get("pcs_set") or "").strip().lower()
    spu_val     = str(payload.get("spu") or "").strip().lower()
    mrp_val     = _safe_float(payload.get("mrp") or 0)
    mrp_min     = _safe_float(payload.get("mrp_min") or 0)
    mrp_max     = _safe_float(payload.get("mrp_max") or 0)
    free_query  = (payload.get("query") or payload.get("item_name") or "").strip().lower()
    ai_enabled  = payload.get("ai", True)
    limit       = int(payload.get("limit") or 20)

    all_products = [p for prods in _FILE_CACHE.values() for p in prods]

    def _multi_score(p: dict) -> int:
        score = 0
        p_code   = (p.get("item_code")  or "").lower()
        p_cat    = (p.get("category")   or "").lower()
        p_size   = (p.get("size")       or "").lower()
        p_finish = (p.get("finish")     or "").lower()
        p_pcs    = (p.get("unit")       or p.get("pcs_/_set") or p.get("pcs / set") or "").lower()
        p_spu    = str(p.get("spu")     or "").lower()
        p_mrp    = _safe_float(p.get("mrp") or 0)
        p_name   = (p.get("item_name")  or "").lower()

        # Item code — strongest signal
        if item_code:
            if item_code == p_code:
                return 100                          # exact code = perfect match
            if p_code.startswith(item_code) or item_code in p_code:
                score += 85

        # Category
        if category:
            if category == p_cat:
                score += 50
            elif category in p_cat or p_cat in category:
                score += 25

        # Size (tokenised match)
        if size:
            size_tokens = set(re.split(r'[\s|x×,]+', size))
            p_size_tokens = set(re.split(r'[\s|x×,]+', p_size))
            common = size_tokens & p_size_tokens - {''}
            if common:
                score += min(40, len(common) * 15)

        # Finish
        if finish:
            if finish in p_finish or p_finish in finish:
                score += 25

        # PCS/Set
        if pcs_set:
            if pcs_set in p_pcs or p_pcs in pcs_set:
                score += 15

        # SPU
        if spu_val and p_spu:
            if spu_val == p_spu:
                score += 20
            elif spu_val in p_spu:
                score += 10

        # MRP exact/range
        if mrp_val > 0 and p_mrp > 0:
            pct_diff = abs(p_mrp - mrp_val) / mrp_val
            if pct_diff < 0.02:
                score += 30
            elif pct_diff < 0.10:
                score += 20
            elif pct_diff < 0.25:
                score += 10
        if mrp_min > 0 and mrp_max > 0 and p_mrp > 0:
            if mrp_min <= p_mrp <= mrp_max:
                score += 25

        # Free-text query
        if free_query:
            q_tokens = set(free_query.split())
            name_tokens = set(p_name.split())
            common_name = q_tokens & name_tokens - {''}
            if common_name:
                score += min(30, len(common_name) * 8)
            elif free_query in p_name or free_query in p_cat:
                score += 15

        return score

    # Score all products
    scored = [{ **p, "_score": _multi_score(p) } for p in all_products if _multi_score(p) > 0]
    scored.sort(key=lambda x: x["_score"], reverse=True)
    top = scored[:limit * 3]

    # Tier results
    exact  = [p for p in top if p["_score"] >= 90][:limit]
    strong = [p for p in top if 55 <= p["_score"] < 90][:limit]
    near   = [p for p in top if 25 <= p["_score"] < 55][:limit]

    # If AI enabled and results are thin, use GPT-4o to enrich/suggest
    if ai_enabled and (not exact and not strong) and (item_code or size or category or free_query):
        combined_query = " ".join(filter(None, [free_query, category, size, finish, item_code]))
        all_hits = [p for p in top if p["_score"] > 0][:20] or scored[:20]
        if all_hits:
            enriched = await _ai_match(combined_query, all_hits)
            strong = enriched[:limit]

    return {
        "exact":  exact,
        "strong": strong,
        "near":   near,
        "total":  len(exact) + len(strong) + len(near),
        "fields_used": {k: v for k, v in {
            "item_code": item_code, "category": category, "size": size,
            "finish": finish, "pcs_per_set": pcs_set, "spu": spu_val,
            "mrp": mrp_val or None, "mrp_range": f"{mrp_min}-{mrp_max}" if mrp_min else None,
            "query": free_query,
        }.items() if v},
    }


@router.post("/product-files/reload")
async def reload_files():
    """Force-reload all product files from disk."""
    global _LOADED
    _LOADED = False
    _FILE_CACHE.clear()
    _ensure_loaded()
    return {
        "loaded_files": list(_FILE_CACHE.keys()),
        "total_products": sum(len(v) for v in _FILE_CACHE.values()),
    }


@router.post("/product-files/ai-build-quote")
async def ai_build_quote_from_requirement(payload: dict):
    """
    Given a natural-language requirement string, search product files and
    return structured line items ready to populate a quotation.
    """
    _ensure_loaded()
    requirement = (payload.get("requirement") or "").strip()
    if not requirement:
        raise HTTPException(400, "requirement text is required")

    # Step 1: keyword extraction via AI (or simple tokenisation fallback)
    keywords = [requirement]
    try:
        from app.core.config import get_settings
        cfg = get_settings()
        if cfg.openai_api_key:
            from openai import AsyncOpenAI
            client = AsyncOpenAI(api_key=cfg.openai_api_key, timeout=30.0)
            resp = await client.chat.completions.create(
                model="gpt-4o",
                messages=[{
                    "role": "user",
                    "content": f"""Extract product search queries from this hardware/sanitary requirement for the Indian market.
Requirement: "{requirement}"
Return JSON: {{"queries": ["query1","query2",...], "line_items": [{{"item_name":"...","quantity":1,"unit":"Nos","category":"..."}},...] }}
Max 5 queries. Line items are the individual products needed.""",
                }],
                max_tokens=400,
                response_format={"type": "json_object"},
            )
            parsed = json.loads(resp.choices[0].message.content)
            keywords = parsed.get("queries", keywords)
            ai_line_items = parsed.get("line_items", [])
        else:
            ai_line_items = []
    except Exception:
        ai_line_items = []

    # Step 2: search for each keyword
    all_results = []
    for kw in keywords[:5]:
        hits = _search_in_list(
            [p for prods in _FILE_CACHE.values() for p in prods],
            kw, "", 5,
        )
        all_results.extend(hits)

    # Dedup
    seen = set()
    deduped = []
    for r in all_results:
        k = r.get("item_code", r.get("item_name", ""))
        if k and k not in seen:
            seen.add(k)
            deduped.append(r)

    # Merge AI-suggested line items with catalog matches
    line_items = []
    for ai_li in ai_line_items:
        name = ai_li.get("item_name", "")
        # Find best catalog match
        match = next((r for r in deduped if name.lower() in (r.get("item_name") or "").lower()), None)
        if match:
            line_items.append({
                "product_name":   match.get("item_name", name),
                "item_code":      match.get("item_code", ""),
                "category":       match.get("category", ai_li.get("category", "")),
                "unit":           match.get("unit", ai_li.get("unit", "Nos")),
                "quantity":       ai_li.get("quantity", 1),
                "unit_price":     _safe_float(match.get("mrp", match.get("mrp_enriched", 0))),
                "hsn_code":       match.get("hsn_code", ""),
                "size":           match.get("size", ""),
                "finish":         match.get("finish", ""),
                "source_file":    match.get("_source_file", ""),
            })
        else:
            line_items.append({
                "product_name": name,
                "item_code":    "",
                "category":     ai_li.get("category", ""),
                "unit":         ai_li.get("unit", "Nos"),
                "quantity":     ai_li.get("quantity", 1),
                "unit_price":   0,
                "source_file":  "AI suggestion",
            })

    # If no AI line items, fall back to top catalog results
    if not line_items:
        for r in deduped[:10]:
            line_items.append({
                "product_name": r.get("item_name", ""),
                "item_code":    r.get("item_code", ""),
                "category":     r.get("category", ""),
                "unit":         r.get("unit", "Nos"),
                "quantity":     1,
                "unit_price":   _safe_float(r.get("mrp", 0)),
                "hsn_code":     r.get("hsn_code", ""),
                "size":         r.get("size", ""),
                "finish":       r.get("finish", ""),
                "source_file":  r.get("_source_file", ""),
            })

    return {
        "requirement": requirement,
        "line_items":  line_items,
        "catalog_hits": deduped[:10],
    }


def _safe_float(v) -> float:
    try:
        return float(str(v).replace(",", "").replace("₹", "").strip() or 0)
    except Exception:
        return 0.0
